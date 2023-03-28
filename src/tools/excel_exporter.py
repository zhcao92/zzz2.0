# -*- coding:utf-8 -*-
import rospy
import rosbag
from openpyxl import *

class ExcelColumn:
    def __init__(self, name):
        self.data = [name]

    def add_data(self, data):
        self.data.append(data)

# we need a can interface so much! or this script has to be changed from time to time.
class Bag2Excel:
    def __init__(self, directory):
        self.bag = rosbag.Bag(directory)
        address_list = directory.split("/")
        address_list.pop()
        self.address = "/".join(address_list)

    def dump(self):
        self.export_vechicle_status()
        self.export_localization_info()

    def export_vechicle_status(self):
        wb = Workbook()
        # ws1 100hz
        ws1 = wb.active
        ws1.title = "Part1"
        result = []
        # IPC_SCU_AccDecelReq 加速度请求
        time = ExcelColumn('时间')
        column = ExcelColumn('加速度请求')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/IPC_SCU_2']):
            time.add_data(t.to_time())
            column.add_data(msg.IPC_SCU_AccDecelReq)
        result.append(time)
        result.append(column)

        # SCU_IPC_SteeringAngleSpd　转向角速度
        column = ExcelColumn('转向角速度')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_1']):
            column.add_data(msg.SCU_IPC_SteeringAngleSpd)
        result.append(column)

        # SCU_IPC_SteeringAngle 转向角度
        column = ExcelColumn('转向角度')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_1']):
            column.add_data(msg.SCU_IPC_SteeringAngle)
        result.append(column)

        # SCU_IPC_FLWheelSpd
        column = ExcelColumn('前左轮速')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_2']):
            column.add_data(msg.SCU_IPC_FLWheelSpd)
        result.append(column)

        # SCU_IPC_FRWheelSpd
        column = ExcelColumn('前右轮速')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_2']):
            column.add_data(msg.SCU_IPC_FRWheelSpd)
        result.append(column)

        # SCU_IPC_RLWheelSpd
        column = ExcelColumn('后左轮速')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_2']):
            column.add_data(msg.SCU_IPC_RLWheelSpd)
        result.append(column)
        # SCU_IPC_RRWheelSpd
        column = ExcelColumn('后右轮速')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_2']):
            column.add_data(msg.SCU_IPC_RRWheelSpd)
        result.append(column)
        # SCU_IPC_dstBat_Dsp
        column = ExcelColumn('续航里程')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_6']):
            column.add_data(msg.SCU_IPC_dstBat_Dsp)
        result.append(column)
        # SCU_IPC_CurrentGearLev
        column = ExcelColumn('当前档位')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_6']):
            column.add_data(msg.SCU_IPC_CurrentGearLev)
        result.append(column)
        # SCU_IPC_AccPedalSig
        column = ExcelColumn('加速踏板信号')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_6']):
            column.add_data(msg.SCU_IPC_AccPedalSig)
        result.append(column)
        # SCU_IPC_BrkPedalSt
        column = ExcelColumn('刹车踏板状态')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_6']):
            column.add_data(msg.SCU_IPC_BrkPedalSt)
        result.append(column)
        # save to sheet 1
        column_number = 1
        for column in result:
            row_number = 1
            for value in column.data:
                ws1.cell(row=row_number, column=column_number).value = value
                row_number += 1
            column_number += 1

        # ws 50hz
        result = []
        ws2 = wb.create_sheet("Part2")
        # SCU_IPC_DBWSt
        time = ExcelColumn('时间')
        column = ExcelColumn('自动驾驶状态')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_3']):
            column.add_data(msg.SCU_IPC_DBWSt)
            time.add_data(t.to_time())
        result.append(time)
        result.append(column)
        # SCU_IPC_VehSpd
        column = ExcelColumn('车辆速度')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_3']):
            column.add_data(msg.SCU_IPC_VehSpd)
        result.append(column)
        # SCU_IPC_BrkPedalSt
        column = ExcelColumn('制动踏板状态')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_3']):
            column.add_data(msg.SCU_IPC_BrkPedalSt)
        result.append(column)
        # SCU_IPC_YAW
        column = ExcelColumn('横摆角速度')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_4']):
            column.add_data(msg.SCU_IPC_YAW)
        result.append(column)
        # SCU_IPC_ActVehLongAccel
        column = ExcelColumn('纵向加速度')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_4']):
            column.add_data(msg.SCU_IPC_ActVehLongAccel)
        result.append(column)
        # SCU_IPC_ActVehLateralAccel
        column = ExcelColumn('横向加速度')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_4']):
            column.add_data(msg.SCU_IPC_ActVehLateralAccel)
        result.append(column)
        # SCU_IPC_EPBSysSt
        column = ExcelColumn('EPB状态')
        for topic, msg, t in self.bag.read_messages(topics=['/canbus/SCU_IPC_4']):
            column.add_data(msg.SCU_IPC_EPBSysSt)
        result.append(column)
        # save to sheet 2
        column_number = 1
        for column in result:
            row_number = 1
            for value in column.data:
                ws2.cell(row=row_number, column=column_number).value = value
                row_number += 1
            column_number += 1

        wb.save(filename=self.address + "/" + "vehicle_state.xlsx")

    def export_localization_info(self):
        result = []
        time = ExcelColumn('时间')
        qx = ExcelColumn('四元数x')
        qy = ExcelColumn('四元数y')
        qz = ExcelColumn('四元数z')
        qw = ExcelColumn('四元数w')
        avx = ExcelColumn('角加速度x')
        avy = ExcelColumn('角加速度y')
        avz = ExcelColumn('角加速度z')
        lax = ExcelColumn('线加速度x')
        lay = ExcelColumn('线加速度x')
        laz = ExcelColumn('线加速度x')
        for topic, msg, t in self.bag.read_messages(topics=['/localization/imu/data']):
            time.add_data(t.to_time())
            qx.add_data(msg.orientation.x)
            qy.add_data(msg.orientation.y)
            qz.add_data(msg.orientation.z)
            qw.add_data(msg.orientation.w)
            avx.add_data(msg.angular_velocity.x)
            avy.add_data(msg.angular_velocity.y)
            avz.add_data(msg.angular_velocity.z)
            lax.add_data(msg.linear_acceleration.x)
            lay.add_data(msg.linear_acceleration.y)
            laz.add_data(msg.linear_acceleration.z)
        result.append(time)
        result.append(qx)
        result.append(qy)
        result.append(qz)
        result.append(qw)
        result.append(avx)
        result.append(avy)
        result.append(avz)
        result.append(lax)
        result.append(lay)
        result.append(laz)

        lat = ExcelColumn('纬度')
        lon = ExcelColumn('经度')
        alt = ExcelColumn('海拔')
        for topic, msg, t in self.bag.read_messages(topics=['/localization/gps/fix']):
            lat.add_data(msg.latitude)
            lon.add_data(msg.longitude)
            alt.add_data(msg.altitude)
        result.append(lat)
        result.append(lon)
        result.append(alt)

        self.dump_to_excel("gps.xlsx", result)

    def dump_to_excel(self, name, data):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "data"
        column_number = 1
        for column in data:
            row_number = 1
            for value in column.data:
                ws1.cell(row=row_number, column=column_number).value = value
                row_number += 1
            column_number += 1
        wb.save(filename=self.address + "/" +name)

if __name__ == "__main__":
    bag = Bag2Excel("/home/icv/zzz/1226morning.bag")
    bag.dump()
